from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from . import services


def _periodo_label(inicio, fim):
    return f'{inicio.strftime("%d/%m/%Y")} a {fim.strftime("%d/%m/%Y")}'


def gerar_excel(inicio, fim):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2D5A8E', end_color='2D5A8E', fill_type='solid')
    header_fill2 = PatternFill(start_color='E11D48', end_color='E11D48', fill_type='solid')
    border = Border(
        bottom=Side(style='thin', color='E2E8F0'),
    )

    def add_header(ws, headers, fill=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if not col_letter and hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # 1. Resumo
    ws = wb.active
    ws.title = 'Resumo'
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Central de Inteligência do Portal'
    ws['A1'].font = Font(bold=True, size=14, color='E11D48')
    ws['A3'] = 'Período:'
    ws['B3'] = _periodo_label(inicio, fim)
    ws['A4'] = 'Gerado em:'
    ws['B4'] = timezone.now().strftime('%d/%m/%Y %H:%M')
    ws['A3'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)

    ind = services.indicadores_resumo(inicio, fim)
    indicadores = [
        ('Usuários online', ind['usuarios_online']),
        ('Visitantes hoje', ind['visitantes_hoje']),
        ('Visitantes semana', ind['visitantes_semana']),
        ('Visitantes mês', ind['visitantes_mes']),
        ('Visitantes ano', ind['visitantes_ano']),
        ('Total geral de acessos', ind['visitantes_total']),
        ('Visitantes únicos', ind['visitantes_unicos']),
        ('Páginas visitadas', ind['paginas_visitadas']),
        ('Downloads', ind['downloads_total']),
        ('Pesquisas', ind['buscas_total']),
        ('Comentários', ind['comentarios_total']),
        ('Conteúdos cadastrados', ind['conteudos_total']),
    ]
    for i, (label, valor) in enumerate(indicadores, 6):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=valor)
    auto_width(ws)

    # 2. Ranking Botões
    ws2 = wb.create_sheet('Ranking Botões')
    add_header(ws2, ['#', 'Botão', 'Acessos', 'Último acesso'])
    for i, item in enumerate(services.ranking_botoes(inicio, fim), 2):
        ws2.cell(row=i, column=1, value=i-1)
        ws2.cell(row=i, column=2, value=item['nome'])
        ws2.cell(row=i, column=3, value=item['total'])
        ws2.cell(row=i, column=4, value=item['ultimo'])
    auto_width(ws2)

    # 3. Ranking Subbotões
    ws3 = wb.create_sheet('Ranking Subbotões')
    add_header(ws3, ['#', 'Subbotão', 'Pai', 'Acessos', 'Último acesso'])
    for i, item in enumerate(services.ranking_subbotoes(inicio, fim), 2):
        ws3.cell(row=i, column=1, value=i-1)
        ws3.cell(row=i, column=2, value=item['nome'])
        ws3.cell(row=i, column=3, value=item['pai'])
        ws3.cell(row=i, column=4, value=item['total'])
        ws3.cell(row=i, column=5, value=item['ultimo'])
    auto_width(ws3)

    # 4. Ranking Documentos
    ws4 = wb.create_sheet('Ranking Documentos')
    add_header(ws4, ['#', 'Documento', 'Tipo', 'Visualizações', 'Último acesso'])
    for i, item in enumerate(services.ranking_documentos(inicio, fim), 2):
        ws4.cell(row=i, column=1, value=i-1)
        ws4.cell(row=i, column=2, value=item['nome'])
        ws4.cell(row=i, column=3, value=item['tipo'])
        ws4.cell(row=i, column=4, value=item['total'])
        ws4.cell(row=i, column=5, value=item['ultimo'])
    auto_width(ws4)

    # 5. Downloads
    ws5 = wb.create_sheet('Downloads')
    add_header(ws5, ['#', 'Arquivo', 'Extensão', 'Downloads'])
    for i, item in enumerate(services.ranking_downloads(inicio, fim, limite=50), 2):
        ws5.cell(row=i, column=1, value=i-1)
        ws5.cell(row=i, column=2, value=item['nome_arquivo'])
        ws5.cell(row=i, column=3, value=item['extensao'])
        ws5.cell(row=i, column=4, value=item['total'])
    auto_width(ws5)

    # 6. Pesquisas
    ws6 = wb.create_sheet('Pesquisas')
    add_header(ws6, ['#', 'Termo pesquisado', 'Pesquisas', 'Última pesquisa'])
    for i, item in enumerate(services.ranking_buscas(inicio, fim, limite=50), 2):
        ws6.cell(row=i, column=1, value=i-1)
        ws6.cell(row=i, column=2, value=item['termo'])
        ws6.cell(row=i, column=3, value=item['total'])
        ws6.cell(row=i, column=4, value=item['ultimo'])
    auto_width(ws6)

    # 7. Comentários
    ws7 = wb.create_sheet('Comentários')
    stats = services.stats_comentarios(inicio, fim)
    add_header(ws7, ['Status', 'Quantidade'])
    for i, (k, v) in enumerate([
        ('Publicados', stats['publicados']),
        ('Pendentes', stats['pendentes']),
        ('Recusados', stats['recusados']),
        ('Respondidos', stats['respondidos']),
        ('Total', stats['total']),
    ], 2):
        ws7.cell(row=i, column=1, value=k)
        ws7.cell(row=i, column=2, value=v)
    auto_width(ws7)

    # 8. Dispositivos e Navegadores
    ws8 = wb.create_sheet('Dispositivos e Navegadores')
    add_header(ws8, ['Dispositivo', 'Acessos', '%'])
    for i, item in enumerate(services.ranking_dispositivos(inicio, fim), 2):
        ws8.cell(row=i, column=1, value=item['nome'])
        ws8.cell(row=i, column=2, value=item['total'])
        ws8.cell(row=i, column=3, value=f"{item['percentual']}%")
    row = ws8.max_row + 2
    ws8.cell(row=row, column=1, value='Navegador').font = Font(bold=True)
    ws8.cell(row=row, column=2, value='Acessos').font = Font(bold=True)
    ws8.cell(row=row, column=3, value='%').font = Font(bold=True)
    for item in services.ranking_navegadores(inicio, fim):
        row += 1
        ws8.cell(row=row, column=1, value=item['nome'])
        ws8.cell(row=row, column=2, value=item['total'])
        ws8.cell(row=row, column=3, value=f"{item['percentual']}%")
    auto_width(ws8)

    # 9. Origem dos acessos
    ws9 = wb.create_sheet('Origem dos Acessos')
    add_header(ws9, ['Origem', 'Acessos', '%'])
    for i, item in enumerate(services.ranking_referrer(inicio, fim), 2):
        ws9.cell(row=i, column=1, value=item['nome'])
        ws9.cell(row=i, column=2, value=item['total'])
        ws9.cell(row=i, column=3, value=f"{item['percentual']}%")
    auto_width(ws9)

    # 10. Áreas do Currículo
    ws10 = wb.create_sheet('Áreas do Currículo')
    add_header(ws10, ['#', 'Área', 'Acessos', '%'])
    for i, item in enumerate(services.ranking_por_area_curricular(inicio, fim), 2):
        ws10.cell(row=i, column=1, value=i-1)
        ws10.cell(row=i, column=2, value=item['nome'])
        ws10.cell(row=i, column=3, value=item['total'])
        ws10.cell(row=i, column=4, value=f"{item['percentual']}%")
    auto_width(ws10)

    # 11. Ranking Banners
    ws11 = wb.create_sheet('Ranking Banners')
    add_header(ws11, ['#', 'Banner', 'Link', 'Impressões'])
    for i, item in enumerate(services.ranking_banners(inicio, fim), 2):
        ws11.cell(row=i, column=1, value=i-1)
        ws11.cell(row=i, column=2, value=item['titulo'])
        ws11.cell(row=i, column=3, value=item['link'])
        ws11.cell(row=i, column=4, value=item['impressoes'])
    auto_width(ws11)

    # 12. Ranking Destaques
    ws12 = wb.create_sheet('Ranking Destaques')
    add_header(ws12, ['#', 'Conteúdo', 'Tipo', 'Impressões', 'Cliques', 'CTR (%)', 'Último'])
    for i, item in enumerate(services.ranking_destaques(inicio, fim), 2):
        ws12.cell(row=i, column=1, value=i-1)
        ws12.cell(row=i, column=2, value=item['titulo'])
        ws12.cell(row=i, column=3, value=item['tipo'])
        ws12.cell(row=i, column=4, value=item['impressoes'])
        ws12.cell(row=i, column=5, value=item['cliques'])
        ws12.cell(row=i, column=6, value=item['ctr'])
        ws12.cell(row=i, column=7, value=item['ultimo'])
    auto_width(ws12)

    # 13. Documentos sem Acesso
    ws13 = wb.create_sheet('Documentos sem Acesso')
    ws13.cell(row=1, column=1, value='Documentos SEM NENHUM acesso').font = Font(bold=True, size=12, color='E11D48')
    row = 3
    add_header_row = ['#', 'Título', 'Tipo', 'Publicado em']

    def _write_secao_docs(ws, row, titulo, docs):
        ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, size=11, color='2D5A8E')
        row += 1
        for col, h in enumerate(add_header_row, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        for i, doc in enumerate(docs, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=doc.get('titulo', ''))
            ws.cell(row=row, column=3, value=doc.get('tipo', ''))
            dp = doc.get('data_publicacao')
            ws.cell(row=row, column=4, value=dp.strftime('%d/%m/%Y') if dp else '')
            row += 1
        return row + 1

    row = _write_secao_docs(ws13, row, 'Nunca acessados', services.documentos_sem_acesso(dias=None))
    row = _write_secao_docs(ws13, row, 'Sem acesso há 30 dias', services.documentos_sem_acesso(dias=30))
    row = _write_secao_docs(ws13, row, 'Sem acesso há 90 dias', services.documentos_sem_acesso(dias=90))
    auto_width(ws13)

    # 14. Alertas
    ws14 = wb.create_sheet('Alertas')
    add_header(ws14, ['#', 'Tipo', 'Título', 'Descrição', 'Data', 'Status'])
    from .models import AlertaInteligencia
    alertas = AlertaInteligencia.objects.all().order_by('-criado_em')[:200]
    for i, a in enumerate(alertas, 2):
        ws14.cell(row=i, column=1, value=i-1)
        ws14.cell(row=i, column=2, value=a.get_tipo_display() if hasattr(a, 'get_tipo_display') else a.tipo)
        ws14.cell(row=i, column=3, value=a.titulo or '')
        ws14.cell(row=i, column=4, value=a.descricao or '')
        ws14.cell(row=i, column=5, value=a.criado_em.strftime('%d/%m/%Y %H:%M') if a.criado_em else '')
        ws14.cell(row=i, column=6, value='Resolvido' if a.resolvido else 'Pendente')
    auto_width(ws14)

    data_str = timezone.now().strftime('%Y-%m-%d')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inteligencia_portal_{data_str}.xlsx"'
    wb.save(response)
    return response


def gerar_pdf(inicio, fim):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=25*mm, bottomMargin=20*mm,
        leftMargin=15*mm, rightMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='IntelTitle',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#e11d48'),
        spaceAfter=6,
        fontName='Helvetica-Bold',
    ))
    styles.add(ParagraphStyle(
        name='IntelSection',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#2d5a8e'),
        spaceBefore=14,
        spaceAfter=6,
        fontName='Helvetica-Bold',
    ))
    styles.add(ParagraphStyle(
        name='IntelNormal',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1e293b'),
    ))

    cor_header = colors.HexColor('#2d5a8e')
    cor_header_text = colors.white
    cor_alt = colors.HexColor('#f8fafc')
    cor_border = colors.HexColor('#e2e8f0')

    table_style_base = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), cor_header),
        ('TEXTCOLOR', (0, 0), (-1, 0), cor_header_text),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, cor_border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, cor_alt]),
    ])

    elements = []

    # Header
    elements.append(Paragraph('Central de Inteligência do Portal', styles['IntelTitle']))
    elements.append(Paragraph(
        f'Período: {_periodo_label(inicio, fim)} &nbsp;&nbsp;|&nbsp;&nbsp; '
        f'Gerado em: {timezone.now().strftime("%d/%m/%Y %H:%M")}',
        styles['IntelNormal']
    ))
    elements.append(Spacer(1, 8*mm))

    # Indicadores
    elements.append(Paragraph('Indicadores Gerais', styles['IntelSection']))
    ind = services.indicadores_resumo(inicio, fim)
    ind_data = [
        ['Indicador', 'Valor'],
        ['Usuários online', str(ind['usuarios_online'])],
        ['Visitantes hoje', str(ind['visitantes_hoje'])],
        ['Visitantes semana', str(ind['visitantes_semana'])],
        ['Visitantes mês', str(ind['visitantes_mes'])],
        ['Visitantes ano', str(ind['visitantes_ano'])],
        ['Total geral', str(ind['visitantes_total'])],
        ['Visitantes únicos (período)', str(ind['visitantes_unicos'])],
        ['Páginas visitadas', str(ind['paginas_visitadas'])],
        ['Downloads', str(ind['downloads_total'])],
        ['Pesquisas', str(ind['buscas_total'])],
        ['Comentários', str(ind['comentarios_total'])],
        ['Conteúdos cadastrados', str(ind['conteudos_total'])],
    ]
    t = Table(ind_data, colWidths=[120*mm, 50*mm])
    t.setStyle(table_style_base)
    elements.append(t)

    # Ranking Botões
    ranking_bot = services.ranking_botoes(inicio, fim)
    if ranking_bot:
        elements.append(Paragraph('Ranking dos Botões', styles['IntelSection']))
        data = [['#', 'Botão', 'Acessos', 'Último']]
        for i, item in enumerate(ranking_bot, 1):
            data.append([str(i), item['nome'], str(item['total']), item['ultimo']])
        t = Table(data, colWidths=[10*mm, 80*mm, 30*mm, 50*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Ranking Subbotões
    ranking_sub = services.ranking_subbotoes(inicio, fim)
    if ranking_sub:
        elements.append(Paragraph('Ranking dos Subbotões', styles['IntelSection']))
        data = [['#', 'Subbotão', 'Pai', 'Acessos', 'Último']]
        for i, item in enumerate(ranking_sub, 1):
            data.append([str(i), item['nome'], item['pai'], str(item['total']), item['ultimo']])
        t = Table(data, colWidths=[10*mm, 65*mm, 45*mm, 25*mm, 35*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Ranking Documentos
    ranking_docs = services.ranking_documentos(inicio, fim)
    if ranking_docs:
        elements.append(Paragraph('Documentos Mais Acessados', styles['IntelSection']))
        data = [['#', 'Documento', 'Tipo', 'Views', 'Último']]
        for i, item in enumerate(ranking_docs, 1):
            titulo = item['nome'][:50] + ('...' if len(item['nome']) > 50 else '')
            data.append([str(i), titulo, item['tipo'], str(item['total']), item['ultimo']])
        t = Table(data, colWidths=[10*mm, 70*mm, 25*mm, 25*mm, 40*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Downloads
    ranking_dl = services.ranking_downloads(inicio, fim)
    if ranking_dl:
        elements.append(Paragraph('Documentos Mais Baixados', styles['IntelSection']))
        data = [['#', 'Arquivo', 'Tipo', 'Downloads']]
        for i, item in enumerate(ranking_dl, 1):
            nome = item['nome_arquivo'][:50] + ('...' if len(item['nome_arquivo']) > 50 else '')
            data.append([str(i), nome, item['extensao'], str(item['total'])])
        t = Table(data, colWidths=[10*mm, 90*mm, 25*mm, 40*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Pesquisas
    ranking_busca = services.ranking_buscas(inicio, fim)
    if ranking_busca:
        elements.append(Paragraph('Pesquisas Mais Realizadas', styles['IntelSection']))
        data = [['#', 'Termo', 'Pesquisas', 'Última']]
        for i, item in enumerate(ranking_busca, 1):
            data.append([str(i), item['termo'], str(item['total']), item['ultimo']])
        t = Table(data, colWidths=[10*mm, 80*mm, 30*mm, 50*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Comentários
    elements.append(Paragraph('Comentários', styles['IntelSection']))
    stats = services.stats_comentarios(inicio, fim)
    data = [
        ['Status', 'Quantidade'],
        ['Publicados', str(stats['publicados'])],
        ['Pendentes', str(stats['pendentes'])],
        ['Recusados', str(stats['recusados'])],
        ['Respondidos', str(stats['respondidos'])],
        ['Total', str(stats['total'])],
    ]
    t = Table(data, colWidths=[120*mm, 50*mm])
    t.setStyle(table_style_base)
    elements.append(t)

    # Dispositivos
    disp = services.ranking_dispositivos(inicio, fim)
    if disp:
        elements.append(Paragraph('Dispositivos', styles['IntelSection']))
        data = [['Dispositivo', 'Acessos', '%']]
        for item in disp:
            data.append([item['nome'], str(item['total']), f"{item['percentual']}%"])
        t = Table(data, colWidths=[80*mm, 45*mm, 40*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Navegadores
    nav = services.ranking_navegadores(inicio, fim)
    if nav:
        elements.append(Paragraph('Navegadores', styles['IntelSection']))
        data = [['Navegador', 'Acessos', '%']]
        for item in nav:
            data.append([item['nome'], str(item['total']), f"{item['percentual']}%"])
        t = Table(data, colWidths=[80*mm, 45*mm, 40*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Origem
    ref = services.ranking_referrer(inicio, fim)
    if ref:
        elements.append(Paragraph('Origem dos Acessos', styles['IntelSection']))
        data = [['Origem', 'Acessos', '%']]
        for item in ref:
            data.append([item['nome'], str(item['total']), f"{item['percentual']}%"])
        t = Table(data, colWidths=[80*mm, 45*mm, 40*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Áreas do currículo
    areas = services.ranking_por_area_curricular(inicio, fim)
    if areas:
        elements.append(Paragraph('Ranking por Área do Currículo', styles['IntelSection']))
        data = [['#', 'Área', 'Acessos', '%']]
        for i, item in enumerate(areas, 1):
            data.append([str(i), item['nome'], str(item['total']), f"{item['percentual']}%"])
        t = Table(data, colWidths=[10*mm, 80*mm, 40*mm, 35*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Ranking Banners
    banners = services.ranking_banners(inicio, fim)
    if banners:
        elements.append(Paragraph('Ranking de Acesso aos Banners', styles['IntelSection']))
        data = [['#', 'Banner', 'Impressões']]
        for i, item in enumerate(banners, 1):
            titulo = item['titulo'][:60] + ('...' if len(item['titulo']) > 60 else '')
            data.append([str(i), titulo, str(item['impressoes'])])
        t = Table(data, colWidths=[10*mm, 120*mm, 35*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Ranking Destaques
    destaques = services.ranking_destaques(inicio, fim)
    if destaques:
        elements.append(Paragraph('Ranking da Área Destaque', styles['IntelSection']))
        data = [['#', 'Conteúdo', 'Tipo', 'Impr.', 'Cliques', 'CTR %']]
        for i, item in enumerate(destaques, 1):
            titulo = item['titulo'][:40] + ('...' if len(item['titulo']) > 40 else '')
            data.append([str(i), titulo, item['tipo'], str(item['impressoes']),
                         str(item['cliques']), f"{item['ctr']}"])
        t = Table(data, colWidths=[10*mm, 65*mm, 30*mm, 20*mm, 25*mm, 20*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Documentos sem acesso
    docs_nunca = services.documentos_sem_acesso(dias=None)
    docs_30 = services.documentos_sem_acesso(dias=30)
    docs_90 = services.documentos_sem_acesso(dias=90)

    def _tabela_docs_sem_acesso(titulo, docs):
        if not docs:
            return
        elements.append(Paragraph(titulo, styles['IntelSection']))
        data = [['#', 'Título', 'Tipo', 'Publicado em']]
        for i, doc in enumerate(docs[:30], 1):
            nome = doc.get('titulo', '')[:50] + ('...' if len(doc.get('titulo', '')) > 50 else '')
            dp = doc.get('data_publicacao')
            data.append([str(i), nome, doc.get('tipo', ''),
                         dp.strftime('%d/%m/%Y') if dp else ''])
        t = Table(data, colWidths=[10*mm, 90*mm, 30*mm, 35*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    _tabela_docs_sem_acesso('Documentos Nunca Acessados', docs_nunca)
    _tabela_docs_sem_acesso('Documentos sem Acesso há 30 dias', docs_30)
    _tabela_docs_sem_acesso('Documentos sem Acesso há 90 dias', docs_90)

    # Alertas
    from .models import AlertaInteligencia
    alertas = list(AlertaInteligencia.objects.all().order_by('-criado_em')[:100])
    if alertas:
        elements.append(Paragraph('Alertas da Central de Inteligência', styles['IntelSection']))
        data = [['#', 'Tipo', 'Título', 'Data', 'Status']]
        for i, a in enumerate(alertas, 1):
            titulo = (a.titulo or '')[:55] + ('...' if len(a.titulo or '') > 55 else '')
            data.append([
                str(i),
                a.get_tipo_display() if hasattr(a, 'get_tipo_display') else a.tipo,
                titulo,
                a.criado_em.strftime('%d/%m/%Y') if a.criado_em else '',
                'Resolvido' if a.resolvido else 'Pendente',
            ])
        t = Table(data, colWidths=[10*mm, 35*mm, 80*mm, 25*mm, 25*mm])
        t.setStyle(table_style_base)
        elements.append(t)

    # Page number callback
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#64748b'))
        canvas.drawCentredString(
            A4[0] / 2, 12*mm,
            f'Central de Inteligência do Portal — Página {doc.page}'
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

    data_str = timezone.now().strftime('%Y-%m-%d')
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inteligencia_portal_{data_str}.pdf"'
    return response
